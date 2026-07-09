"""Тесты серверного .xlsx-экспорта — офлайн, без БД (build_xlsx над Incident)."""

import zipfile
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import load_workbook

from app.models import Incident
from app.services.export import build_xlsx

# 0-based индексы колонок. Столбец «Ссылка на фото» заменён на 3 столбца «Фото 1/2/3»
# (формула =IMAGE(url)) → «Ссылка на сообщение»/«Поступило» сдвинуты на +2.
_COMMENT_COL = 8  # «Комментарий» (сразу после «Координаты»)
_PHOTO1_COL = 12  # «Фото 1»
_PHOTO2_COL = 13  # «Фото 2»
_PHOTO3_COL = 14  # «Фото 3»
_MSG_LINK_COL = 15  # «Ссылка на сообщение»


def _incident(**kw) -> Incident:
    base = dict(
        source="max",
        status="new",
        fio="Иванов Иван",
        region="Самарская обл",
        city="г Кинель",
        street="ул Маяковского, д 41",
        coords="53.2, 50.6",
        photo_time=datetime(2026, 4, 26, 8, 5, tzinfo=timezone.utc),
        photos=1,
        msg="mid.ffff0001",
        msg_url=None,
        received_at=datetime(2026, 4, 26, 8, 11, tzinfo=timezone.utc),
    )
    base.update(kw)
    return Incident(**base)


def _link_cells(rows: list[Incident]) -> list:
    """Возвращает значения колонки «Ссылка на сообщение» (без строки заголовков)."""
    wb = load_workbook(BytesIO(build_xlsx(rows)))
    ws = wb.active
    return [row[_MSG_LINK_COL].value for row in ws.iter_rows(min_row=2)]


def test_export_writes_msg_url_as_is():
    """msg_url непустой → попадает в колонку «Ссылка на сообщение» КАК ЕСТЬ."""
    url = "https://max.ru/c/-75787158905457/AZ8DNeZnbkM"
    inc = _incident(msg_url=url)
    assert _link_cells([inc]) == [url]


def test_export_blank_msg_url_writes_empty_not_mid_link():
    """msg_url пуст → пустая строка в колонке; БИТАЯ https://max.ru/m/{mid} НЕ строится."""
    inc = _incident(msg="mid.ffffdead", msg_url=None)
    cells = _link_cells([inc])
    # Колонка пустая (openpyxl отдаёт None для "") — главное: НЕ ссылка на mid.
    assert cells[0] in (None, "")
    assert "max.ru/m/" not in (cells[0] or "")
    assert "mid.ffffdead" not in (cells[0] or "")


def _photo_cell(rows: list[Incident], col: int, base_url: str = ""):
    """Ячейка указанного столбца «Фото N» первой строки данных (для hyperlink/value)."""
    wb = load_workbook(BytesIO(build_xlsx(rows, base_url)))
    ws = wb.active
    return next(ws.iter_rows(min_row=2))[col]


def test_export_photo_cell_has_hyperlink_to_full():
    """Столбцы «Фото 1/2/3» — гиперссылка на ПОЛНЫЙ URL фото (клик → скачать); нет фото → пусто.

    Файла на диске в тесте нет → картинка не встраивается, ячейка = кликабельный текст
    «Открыть фото» с гиперссылкой (в проде вместо текста будет встроенная миниатюра)."""
    inc = _incident(
        photos=2,
        photo_urls=[
            "/api/v1/intake/photo/abc/0.jpg",
            "/api/v1/intake/photo/abc/1.jpg",
        ],
    )
    c1 = _photo_cell([inc], _PHOTO1_COL, "https://ecopulse.reo.ru")
    c2 = _photo_cell([inc], _PHOTO2_COL, "https://ecopulse.reo.ru")
    c3 = _photo_cell([inc], _PHOTO3_COL, "https://ecopulse.reo.ru")
    assert c1.hyperlink.target == "https://ecopulse.reo.ru/api/v1/intake/photo/abc/0.jpg"
    assert c2.hyperlink.target == "https://ecopulse.reo.ru/api/v1/intake/photo/abc/1.jpg"
    assert c1.value == "Открыть фото"  # без файла — текстовая ссылка
    assert c3.hyperlink is None and c3.value in (None, "")  # третьего фото нет


def test_export_embeds_thumb_and_keeps_link(tmp_path, monkeypatch):
    """Есть миниатюра на диске → картинка ВСТРОЕНА в файл (xl/media/*) + ячейка с гиперссылкой."""
    from PIL import Image as PILImage

    monkeypatch.setattr("app.services.export.settings.STORAGE_DIR", str(tmp_path))
    inc_id = "abc123"
    photo_dir = tmp_path / "incidents" / inc_id
    photo_dir.mkdir(parents=True)
    PILImage.new("RGB", (120, 90), "green").save(photo_dir / "0_thumb.jpg")

    inc = _incident(photos=1, photo_urls=[f"/api/v1/intake/photo/{inc_id}/0.jpg"])
    data = build_xlsx([inc], "https://ecopulse.reo.ru")

    # Картинка реально встроена в книгу (медиа внутри xlsx-архива).
    with zipfile.ZipFile(BytesIO(data)) as z:
        assert any(n.startswith("xl/media/") for n in z.namelist())
    # Гиперссылка на полное фото сохранена на ячейке «Фото 1».
    wb = load_workbook(BytesIO(data))
    cell = next(wb.active.iter_rows(min_row=2))[_PHOTO1_COL]
    assert cell.hyperlink.target == f"https://ecopulse.reo.ru/api/v1/intake/photo/{inc_id}/0.jpg"


def test_export_photo_skip_placeholder():
    """Плейсхолдеры демо-сида (placeholder://…) → пустая ячейка (нет URL — нет ссылки/картинки)."""
    inc = _incident(photos=1, photo_urls=["placeholder://incident-photo/1"])
    c = _photo_cell([inc], _PHOTO1_COL, "https://ecopulse.reo.ru")
    assert c.hyperlink is None and c.value in (None, "")


def test_export_has_photo_headers():
    """Заголовки «Фото 1/2/3» на месте, перед «Ссылка на сообщение»."""
    wb = load_workbook(BytesIO(build_xlsx([])))
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert headers[_PHOTO1_COL] == "Фото 1"
    assert headers[_PHOTO2_COL] == "Фото 2"
    assert headers[_PHOTO3_COL] == "Фото 3"
    assert headers[_MSG_LINK_COL] == "Ссылка на сообщение"


def _comment_cells(rows: list[Incident]) -> list:
    """Значения колонки «Комментарий» (без строки заголовков)."""
    wb = load_workbook(BytesIO(build_xlsx(rows)))
    ws = wb.active
    return [row[_COMMENT_COL].value for row in ws.iter_rows(min_row=2)]


def test_export_has_comment_header():
    """Заголовок «Комментарий» на месте — сразу после «Координаты»."""
    wb = load_workbook(BytesIO(build_xlsx([])))
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert headers[_COMMENT_COL] == "Комментарий"
    assert headers[_COMMENT_COL - 1] == "Координаты"


def test_export_writes_comment_value():
    """Непустой comment попадает в колонку «Комментарий» как есть."""
    inc = _incident(comment="Радар №116434; Баки раздельного сбора отсутствуют")
    assert _comment_cells([inc]) == [
        "Радар №116434; Баки раздельного сбора отсутствуют"
    ]


def test_export_blank_comment_writes_empty():
    """comment=None → пустая колонка (openpyxl отдаёт None для '')."""
    inc = _incident(comment=None)
    assert _comment_cells([inc])[0] in (None, "")
