"""Тесты выгрузки в формате УТКО (build_utko_xlsx) — офлайн, без БД."""

from datetime import datetime, timezone
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import load_workbook

from app.models import Incident
from app.services import region as region_service
from app.services.addr_norm import region_match_key
from app.services.utko_export import build_utko_xlsx


def _incident(**kw) -> Incident:
    base = dict(
        source="form",
        status="new",
        fio="Иванов Иван",
        region="Самарская область",
        city="г. Кинель",
        street="ул. Маяковского, 41",
        coords="53.2, 50.6",
        mno_reg="63-04-001162",
        incident_type="fire",
        comment="Возгорание в контейнере",
        photo_time=datetime(2026, 4, 26, 8, 5, tzinfo=timezone.utc),
        photos=2,
        photo_urls=[
            "/api/v1/intake/photo/abc/0.jpg",
            "/api/v1/intake/photo/abc/1.jpg",
        ],
        received_at=datetime(2026, 4, 26, 8, 11, tzinfo=timezone.utc),
    )
    base.update(kw)
    return Incident(**base)


_EXPECTED_HEADERS = [
    "Субъект РФ",
    "Реестровый номер МНО",
    "Дата и время фотофиксации",
    "Адрес",
    "Тип инцидента",
    "Подтип инцидента",
    "Описание",
    "Ссылка на фото",
    "Ссылка на фото",
    "Ссылка на фото",
    "Ссылка на фото",
    "Ссылка на фото",
    "Ссылка на фото",
]


def _rows(rows, base_url=""):
    wb = load_workbook(BytesIO(build_utko_xlsx(rows, base_url, {"fire": "Возгорание в контейнере"})))
    return wb.active


def test_utko_headers_exact_order():
    ws = _rows([])
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert headers == _EXPECTED_HEADERS


def test_utko_template_header_intact_data_from_row_5():
    """Шапка шаблона (строки 1–4) не трогается, данные — строго с 5-й строки."""
    inc = _incident()
    ws = _rows([inc])
    # лист данных из шаблона
    assert ws.title == "Реестр для инцидентов"
    # строки 2–4 — фиксированная шапка шаблона (гвоздями): номера · типы · обязательность
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=3, column=1).value == "Список"
    assert ws.cell(row=4, column=1).value == "Обязательное поле"
    # первая строка данных — 5-я
    assert ws.cell(row=5, column=1).value == "Самарская область"
    assert ws.cell(row=5, column=2).value == "63-04-001162"


def test_utko_row_maps_fields_and_photo_urls():
    inc = _incident()
    ws = _rows([inc], base_url="https://ecopulse.reo.ru")
    row = [c.value for c in next(ws.iter_rows(min_row=5))]
    assert row[0] == "Самарская область"          # Субъект РФ
    assert row[1] == "63-04-001162"               # Реестровый номер МНО
    assert row[2] == "26.04.2026 08:05"           # Дата и время фотофиксации
    assert row[3] == "г. Кинель, ул. Маяковского, 41"  # Адрес (город + улица)
    assert row[4] == "Возгорание в контейнере"    # Тип инцидента (по type_labels)
    assert row[5] in (None, "")                    # Подтип — пусто (поля нет)
    assert row[6] == "Возгорание в контейнере"    # Описание (comment)
    # 6 столбцов «Ссылка на фото»: 2 полных URL + 4 пустых.
    assert row[7] == "https://ecopulse.reo.ru/api/v1/intake/photo/abc/0.jpg"
    assert row[8] == "https://ecopulse.reo.ru/api/v1/intake/photo/abc/1.jpg"
    assert all(row[i] in (None, "") for i in range(9, 13))


def test_utko_subject_from_our_directory_by_mno():
    """«Субъект РФ» берётся из справочника (region_by_mno по mno_id), а не из inc.region (DaData)."""
    from uuid import uuid4

    mid = uuid4()
    inc = _incident(region="Самарская обл (DaData)", mno_id=mid)
    wb = load_workbook(
        BytesIO(
            build_utko_xlsx([inc], "", {"fire": "x"}, region_by_mno={mid: "Самарская область"})
        )
    )
    row = [c.value for c in next(wb.active.iter_rows(min_row=5))]
    assert row[0] == "Самарская область"  # из нашего справочника (как в УТКО), НЕ DaData

    # Нет соответствия в справочнике / нет карты → фолбэк на inc.region.
    wb2 = load_workbook(BytesIO(build_utko_xlsx([inc], "", {"fire": "x"}, region_by_mno={})))
    row2 = [c.value for c in next(wb2.active.iter_rows(min_row=5))]
    assert row2[0] == "Самарская обл (DaData)"


# --- «Субъект РФ» из справочника для инцидентов БЕЗ МНО ------------------------

# 85 субъектов РОВНО как их отдаёт ФГИС (GET filters/regions) и как они записаны в листе
# «Субъекты РФ» шаблона УТКО — символ в символ (обычный дефис в «… - Югра», «г. Москва»).
# Именно эти имена лежат в нашей таблице regions (синхр. из ФГИС). Зафиксированы копией:
# тесты офлайн, в сеть НЕ ходят. (В шаблоне есть ещё 5 — новые территории и Байконур:
# ФГИС их не отдаёт, в справочнике их нет.)
_FGIS_SUBJECTS = [
    "Республика Адыгея", "Республика Башкортостан", "Республика Бурятия",
    "Республика Алтай", "Республика Дагестан", "Республика Ингушетия",
    "Кабардино-Балкарская Республика", "Республика Калмыкия",
    "Карачаево-Черкесская Республика", "Республика Карелия", "Республика Коми",
    "Республика Марий Эл", "Республика Мордовия", "Республика Саха (Якутия)",
    "Республика Северная Осетия - Алания", "Республика Татарстан", "Республика Тыва",
    "Удмуртская Республика", "Республика Хакасия", "Чеченская Республика",
    "Чувашская Республика - Чувашия", "Алтайский край", "Краснодарский край",
    "Красноярский край", "Приморский край", "Ставропольский край", "Хабаровский край",
    "Амурская область", "Архангельская область", "Астраханская область",
    "Белгородская область", "Брянская область", "Владимирская область",
    "Волгоградская область", "Вологодская область", "Воронежская область",
    "Ивановская область", "Иркутская область", "Калининградская область",
    "Калужская область", "Камчатский край", "Кемеровская область - Кузбасс",
    "Кировская область", "Костромская область", "Курганская область", "Курская область",
    "Ленинградская область", "Липецкая область", "Магаданская область",
    "Московская область", "Мурманская область", "Нижегородская область",
    "Новгородская область", "Новосибирская область", "Омская область",
    "Оренбургская область", "Орловская область", "Пензенская область", "Пермский край",
    "Псковская область", "Ростовская область", "Рязанская область", "Самарская область",
    "Саратовская область", "Сахалинская область", "Свердловская область",
    "Смоленская область", "Тамбовская область", "Тверская область", "Томская область",
    "Тульская область", "Тюменская область", "Ульяновская область",
    "Челябинская область", "Забайкальский край", "Ярославская область", "г. Москва",
    "г. Санкт-Петербург", "Еврейская автономная область", "Ненецкий автономный округ",
    "Ханты-Мансийский автономный округ - Югра", "Чукотский автономный округ",
    "Ямало-Ненецкий автономный округ", "Республика Крым", "г. Севастополь",
]


async def _canonical_index(names=_FGIS_SUBJECTS) -> dict[str, str]:
    """Реальный region.canonical_index на поддельной сессии (БД в тестах нет).

    Формула ключа в тесте НЕ дублируется — индекс строит production-код.
    """
    session = AsyncMock()
    session.execute.return_value = MagicMock(
        scalars=MagicMock(return_value=MagicMock(all=MagicMock(return_value=list(names))))
    )
    return await region_service.canonical_index(session)


async def _subject_of(region: str, **kw) -> str:
    """«Субъект РФ» (колонка 1) выгрузки инцидента с данным inc.region."""
    inc = _incident(region=region, **kw)
    index = await _canonical_index()
    wb = load_workbook(
        BytesIO(build_utko_xlsx([inc], "", {"fire": "x"}, region_index=index))
    )
    return next(wb.active.iter_rows(min_row=5))[0].value


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "raw, expected",
    [
        # Главный кейс жалобы: normalize_region срезает ведущий «г.» → в inc.region лежит
        # «Санкт-Петербург», а УТКО принимает только «г. Санкт-Петербург».
        ("Санкт-Петербург", "г. Санкт-Петербург"),
        ("Москва", "г. Москва"),
        ("Севастополь", "г. Севастополь"),
        # Сокращённые формы DaData.
        ("Респ Татарстан", "Республика Татарстан"),
        ("Мурманская обл", "Мурманская область"),
        ("Еврейская Аобл", "Еврейская автономная область"),
        # Длинное тире (DaData/AI) → канон справочника с обычным дефисом.
        ("Ханты-Мансийский автономный округ — Югра", "Ханты-Мансийский автономный округ - Югра"),
        ("Кемеровская область—Кузбасс", "Кемеровская область - Кузбасс"),
        # Регистр и «ё».
        ("САНКТ-ПЕТЕРБУРГ", "г. Санкт-Петербург"),
        ("самарская область", "Самарская область"),
    ],
)
async def test_utko_subject_resolved_from_directory_without_mno(raw, expected):
    """Инцидент БЕЗ МНО: inc.region сопоставляется со справочником → каноническое имя."""
    assert await _subject_of(raw) == expected


@pytest.mark.asyncio
async def test_utko_subject_mno_wins_over_incident_region_text():
    """МНО из ФГИС авторитетнее текста inc.region: справочник по mno_id имеет приоритет."""
    from uuid import uuid4

    mid = uuid4()
    inc = _incident(region="Москва", mno_id=mid)  # текст резолвился бы в «г. Москва»
    index = await _canonical_index()
    wb = load_workbook(
        BytesIO(
            build_utko_xlsx(
                [inc], "", {"fire": "x"},
                region_by_mno={mid: "Самарская область"},
                region_index=index,
            )
        )
    )
    row = [c.value for c in next(wb.active.iter_rows(min_row=5))]
    assert row[0] == "Самарская область"  # из МНО, а не из inc.region


@pytest.mark.asyncio
async def test_utko_subject_unknown_region_kept_as_is():
    """Субъекта нет в справочнике → пишем текст как есть (данные не теряем)."""
    assert await _subject_of("Нарния") == "Нарния"


@pytest.mark.asyncio
async def test_utko_subject_empty_region_does_not_crash():
    """Пустой inc.region не роняет выгрузку и не подхватывает случайный канон."""
    assert await _subject_of("") in (None, "")


@pytest.mark.asyncio
async def test_region_match_key_no_collisions_on_all_85_subjects():
    """Ключ сопоставления различает все 85 субъектов и резолвит каждого в себя."""
    index = await _canonical_index()
    assert len(_FGIS_SUBJECTS) == 85
    assert len(index) == 85  # ключи уникальны — ни один субъект не затёр другого
    assert len({region_match_key(n) for n in _FGIS_SUBJECTS}) == 85
    for name in _FGIS_SUBJECTS:
        assert await _subject_of(name) == name


def test_utko_subtype_label_for_no_access():
    """incident_subtype='blocked_by_car' → в колонке «Подтип инцидента» его подпись."""
    inc = _incident(incident_type="no_access", incident_subtype="blocked_by_car")
    ws = _rows([inc])
    row = [c.value for c in next(ws.iter_rows(min_row=5))]
    assert row[5] == "Проезд заблокирован автомобилем"  # Подтип из справочника


def test_utko_subtype_empty_when_absent():
    """Без подтипа (incident_subtype=None) колонка «Подтип инцидента» пуста."""
    inc = _incident(incident_subtype=None)
    ws = _rows([inc])
    row = [c.value for c in next(ws.iter_rows(min_row=5))]
    assert row[5] in (None, "")


def test_utko_skips_placeholder_photos():
    inc = _incident(photos=1, photo_urls=["placeholder://incident-photo/1"])
    ws = _rows([inc], base_url="https://ecopulse.reo.ru")
    row = [c.value for c in next(ws.iter_rows(min_row=5))]
    assert all(row[i] in (None, "") for i in range(7, 13))  # плейсхолдер → все ссылки пусты
