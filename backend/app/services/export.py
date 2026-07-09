"""Серверный экспорт инцидентов в .xlsx (openpyxl). 17 колонок: порядок ТЗ §7 + комментарий +
3 столбца «Фото 1/2/3». В каждый столбец ВСТРАИВАЕТСЯ превью фото (миниатюра с диска) — видно
в ЛЮБОМ Excel/МойОфис/Р7/LibreOffice, без интернета и без макросов; ячейка при этом несёт
гиперссылку на ПОЛНОЕ фото (клик → скачать полноразмерное, нужен интернет)."""

import logging
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from ..config import settings
from ..models import Incident

logger = logging.getLogger(__name__)

# Превью фото в ячейке: максимальная сторона миниатюры (px), высота строки (pt) и ширина
# столбцов «Фото» — с запасом, чтобы картинка была видна, а по краю ячейки работала
# гиперссылка (клик → полное фото).
_IMG_MAX_PX = 90
_ROW_HEIGHT_PT = 74  # ≈ 98px — вмещает картинку 90px + поле
_PHOTO_COL_WIDTH = 14  # ≈ 100px

# Лейблы — зеркало data.js STATUS/SOURCE
_STATUS_LABELS = {
    "new": "Новый",
    "found": "Инцидент обнаружен",
    "none": "Нет инцидента",
    "exported": "Выгружен",
}
_SOURCE_LABELS = {
    "max": "Макс",
    "form": "Яндекс форма",
}

# Заголовки в точном порядке (ТЗ §7 / прототип toCsv)
_HEADERS = [
    "Заявитель",
    "Источник",
    "Статус",
    "Регион",
    "Город / н.п.",
    "Адрес (улица)",
    "Полный адрес",
    "Координаты",
    "Комментарий",
    "Дата фотофиксации",
    "Время фотофиксации",
    "Кол-во фото",
    "Фото 1",
    "Фото 2",
    "Фото 3",
    "Ссылка на сообщение",
    "Поступило",
]

# 1-based номера столбцов «Фото 1/2/3» (для ширины колонок в build_xlsx). Совпадают с
# позицией в _HEADERS: … Кол-во фото(12) · Фото1(13) · Фото2(14) · Фото3(15) · …
_PHOTO_COLS = (13, 14, 15)


def _full_addr(inc: Incident) -> str:
    return f"{inc.region}, {inc.city}, {inc.street}"


def _photo_date(inc: Incident) -> str:
    """ДД.ММ.ГГГГ или '' если фотофиксации нет."""
    if inc.photo_time is None:
        return ""
    return inc.photo_time.strftime("%d.%m.%Y")


def _photo_time(inc: Incident) -> str:
    """ЧЧ:ММ или '' если фотофиксации нет."""
    if inc.photo_time is None:
        return ""
    return inc.photo_time.strftime("%H:%M")


def _message_link(inc: Incident) -> str:
    """Готовый https-URL сообщения (msg_url) как есть; иначе '' (без фолбэка на mid)."""
    return inc.msg_url or ""


def _abs_photo_url(inc: Incident, base_url: str, idx: int) -> str:
    """Полный публичный URL фото №idx (0-based) или '' если фото нет.

    photo_urls хранит относительные пути /api/v1/intake/photo/{id}/{i}.jpg → добавляем
    base_url (схема+домен). Плейсхолдеры демо-сида (placeholder://…) — реального файла нет,
    возвращаем ''. Эндпоинт отдачи фото публичный, URL открывается без авторизации.
    """
    urls = inc.photo_urls or []
    if idx >= len(urls):
        return ""
    u = urls[idx]
    if not isinstance(u, str) or not u:
        return ""
    if u.startswith("http"):
        return u
    if u.startswith("/"):
        return f"{base_url.rstrip('/')}{u}"
    return ""


def _photo_thumb_path(inc: Incident, idx: int) -> Path | None:
    """Путь к файлу-миниатюре фото №idx на диске (или None).

    photo_urls[idx] = /api/v1/intake/photo/{id}/{i}.jpg → на диске лежит миниатюра
    {STORAGE_DIR}/incidents/{id}/{i}_thumb.jpg (её и встраиваем — она лёгкая). Если
    миниатюры нет — пробуем полный {i}.jpg. Плейсхолдеры сида (placeholder://…) и
    отсутствующие файлы → None (встраивать нечего, останется текстовая гиперссылка).
    """
    urls = inc.photo_urls or []
    if idx >= len(urls):
        return None
    u = urls[idx]
    if not isinstance(u, str) or "/intake/photo/" not in u:
        return None
    tail = u.split("/intake/photo/", 1)[1]  # {id}/{i}.jpg
    parts = tail.split("/")
    if len(parts) < 2:
        return None
    ident, filename = parts[-2], parts[-1]
    stem = filename.rsplit(".", 1)[0]
    base = Path(settings.STORAGE_DIR) / "incidents" / ident
    thumb = base / f"{stem}_thumb.jpg"
    if thumb.exists():
        return thumb
    full = base / filename
    return full if full.exists() else None


def _received(inc: Incident) -> str:
    if inc.received_at is None:
        return ""
    return inc.received_at.strftime("%Y-%m-%d %H:%M")


def _row(inc: Incident, base_url: str) -> list:
    return [
        inc.fio,
        _SOURCE_LABELS.get(inc.source, inc.source),
        _STATUS_LABELS.get(inc.status, inc.status),
        inc.region,
        inc.city,
        inc.street,
        _full_addr(inc),
        inc.coords,
        inc.comment or "",
        _photo_date(inc),
        _photo_time(inc),
        inc.photos,
        # Столбцы «Фото 1/2/3» наполняются в build_xlsx (встраивание картинки +
        # гиперссылка на ячейке) — здесь плейсхолдеры, чтобы не сбить порядок колонок.
        "",
        "",
        "",
        _message_link(inc),
        _received(inc),
    ]


def _place_photo(ws, inc: Incident, base_url: str, row_idx: int, idx: int, col: int) -> bool:
    """Ставит в ячейку (row_idx, col) превью фото №idx + гиперссылку на полное фото.

    Возвращает True, если что-то положено (для подъёма высоты строки). Ссылка на ячейке
    работает в любом Excel (клик по краю ячейки → скачать полное фото по URL). Встроенная
    картинка (миниатюра с диска) видна и БЕЗ интернета; битый/отсутствующий файл — не рушит
    экспорт, останется одна текстовая гиперссылка «Открыть фото».
    """
    url = _abs_photo_url(inc, base_url, idx)
    if not url:
        return False
    cell = ws.cell(row=row_idx, column=col)
    cell.hyperlink = url  # клик → полноразмерное фото (нужен интернет)
    thumb = _photo_thumb_path(inc, idx)
    if thumb is not None:
        try:
            img = XLImage(str(thumb))
            longest = max(img.width or _IMG_MAX_PX, img.height or _IMG_MAX_PX)
            scale = _IMG_MAX_PX / longest
            if scale < 1:
                img.width = int(img.width * scale)
                img.height = int(img.height * scale)
            img.anchor = f"{get_column_letter(col)}{row_idx}"
            ws.add_image(img)
            return True
        except Exception:  # noqa: BLE001 — битый файл не должен ронять весь отчёт
            logger.warning("export: не удалось встроить фото %s", thumb, exc_info=True)
    # Картинки нет — оставляем кликабельный текст, чтобы фото всё равно можно было открыть.
    cell.value = "Открыть фото"
    return True


def build_xlsx(rows: Iterable[Incident], base_url: str = "") -> bytes:
    """Строит .xlsx (bytes) из инцидентов. 17 колонок, первая строка — заголовки.

    base_url (схема+домен, напр. https://ecopulse.reo.ru) — для абсолютных URL фото.
    Столбцы «Фото 1/2/3»: в ячейку ВСТРАИВАЕТСЯ миниатюра (видна в любом Excel/МойОфис,
    без интернета) + гиперссылка на полное фото (клик → скачать, нужен интернет). Строки
    с фото делаем выше, столбцы «Фото» шире — под превью.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Инциденты"

    ws.append(_HEADERS)
    row_idx = 1
    for inc in rows:
        row_idx += 1
        ws.append(_row(inc, base_url))
        has_photo = False
        for i, col in enumerate(_PHOTO_COLS):
            if _place_photo(ws, inc, base_url, row_idx, i, col):
                has_photo = True
        if has_photo:
            ws.row_dimensions[row_idx].height = _ROW_HEIGHT_PT

    for col in _PHOTO_COLS:
        ws.column_dimensions[get_column_letter(col)].width = _PHOTO_COL_WIDTH

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
