"""Выгрузка обращений в формате ФГИС УТКО (.xlsx) — СТРОГО по шаблону клиента.

Заполняется ГОТОВЫЙ шаблон `app/templates/utko_template.xlsx` (=клиентский shp.xlsx):
лист «Реестр для инцидентов» имеет ФИКСИРОВАННУЮ 4-строчную шапку (заголовки / номера
колонок / типы данных / «Обязательное поле»), плюс листы-справочники (Субъекты РФ, Тип/
Подтип инцидента, Да_Нет). Наши данные пишутся СТРОГО С 5-Й СТРОКИ — шапку не трогаем
(«гвоздями вбитый» шаблон). Порядок 13 колонок совпадает со строкой 1 шаблона: Субъект РФ ·
Реестровый № МНО · Дата и время фотофиксации · Адрес · Тип · Подтип · Описание · Ссылка
на фото ×6 (просто текстовые полные URL). «Подтип» — подпись из services/incident_subtypes.

«Дата и время фотофиксации» отдаётся в UTC: photo_time хранится как МЕСТНОЕ настенное
время волонтёра (services/intake._photo_wallclock — смещение отброшено, значение помечено
UTC), а УТКО читает поле как UTC, поэтому пересчитываем настенное время по часовому поясу
РЕГИОНА инцидента (services/region_tz.region_utc_offset). Пояс резолвится по коду субъекта
инцидента — ЗЕРКАЛО «Субъекта РФ» (сперва по МНО, иначе по тексту inc.region; см.
_region_code). ⚠️ обычный export.py (админка) время НЕ пересчитывает — там оно местное.

Ссылки на фото ведут на УТКО-КОПИЮ `{i}_utko.jpg` (≤1280×720 и ≤500 КБ), а НЕ на FULL
`{i}.jpg`: ФГИС УТКО отвергает фото >500 КБ, а full-копия админки бывает крупнее. URL
берётся из inc.photo_urls (`.../photo/{id}/{i}.jpg`) вставкой суффикса `_utko` (эндпоинт
get_photo лениво сгенерит копию из FULL, если её ещё нет на диске).

«Субъект РФ» ВСЕГДА резолвится в имя из НАШЕГО справочника регионов (синхр. из ФГИС —
там формы, которые УТКО принимает: «г. Санкт-Петербург», «Кемеровская область - Кузбасс»):
сперва по МНО инцидента, иначе — сопоставлением текста inc.region со справочником
(см. _subject). Сырой DaData-текст в файл уходит только если субъекта нет в справочнике.

⚠️ openpyxl при загрузке/сохранении удаляет расширенные data-validation (выпадашки x14)
шаблона — на СОДЕРЖИМОЕ загрузки в УТКО это не влияет (грузятся значения строк 5+), но
выпадающие списки в выходном файле пропадают. Сам шаблон в репозитории их сохраняет.
"""

import re
from datetime import timedelta
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from ..models import Incident
from .addr_norm import normalize_region, region_match_key
from .incident_subtypes import label_for as _subtype_label
from .incident_types import incident_type_label as _static_type_label
from .region_tz import region_utc_offset

# Готовый шаблон УТКО (клиентский). Лежит в образе (COPY . . в backend/Dockerfile).
_TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "utko_template.xlsx"
# Лист с данными и первая строка данных (1–4 — фиксированная шапка шаблона).
_DATA_SHEET = "Реестр для инцидентов"
_DATA_START_ROW = 5
_PHOTO_SLOTS = 6  # столбцов «Ссылка на фото»


def _photo_datetime(inc: Incident, offset: int) -> str:
    """«ДД.ММ.ГГГГ ЧЧ:ММ» в UTC или '' если фотофиксации нет.

    photo_time хранится как МЕСТНОЕ настенное время волонтёра (services/intake.
    _photo_wallclock — исходное смещение отброшено, значение помечено UTC). УТКО же
    читает поле как UTC, поэтому вычитаем часовой пояс региона инцидента (offset,
    services/region_tz): настенное − offset = реальный UTC. Дата может сдвинуться
    через полночь — это корректно (реальный момент времени тот же).
    """
    if inc.photo_time is None:
        return ""
    utc_dt = inc.photo_time - timedelta(hours=offset)
    return utc_dt.strftime("%d.%m.%Y %H:%M")


def _address(inc: Incident) -> str:
    """Адрес без субъекта РФ (он отдельной колонкой): город + улица."""
    return ", ".join(p for p in [inc.city, inc.street] if p)


def _type_label(inc: Incident, type_labels: dict | None) -> str:
    code = inc.incident_type
    if not code:
        return ""
    if type_labels and code in type_labels:
        return type_labels[code]
    return _static_type_label(code) or code


def _abs_url(u, base_url: str) -> str:
    """Полный публичный URL фото или '' (плейсхолдеры сида/мусор → '')."""
    if not isinstance(u, str) or not u:
        return ""
    if u.startswith("http"):
        return u
    if u.startswith("/"):
        return f"{base_url.rstrip('/')}{u}"
    return ""


def _utko_variant_url(u: str) -> str:
    """URL нашего фото инцидента → ссылка на УТКО-копию `{i}_utko.jpg`.

    Вставляет суффикс `_utko` перед финальным `.jpg`, ТОЛЬКО если это наш URL фото
    инцидента (путь `.../api/v1/intake/photo/{id}/{i}.jpg`, относительный или полный).
    Прочие/внешние URL/плейсхолдеры не трогаем — вернём как есть.
    """
    if re.search(r"/api/v1/intake/photo/[^/]+/[0-9]+\.jpg$", u):
        return u[: -len(".jpg")] + "_utko.jpg"
    return u


def _photo_urls(inc: Incident, base_url: str) -> list[str]:
    """До _PHOTO_SLOTS полных URL УТКО-копий фото; недостающие — пустые (ровно _PHOTO_SLOTS).

    Ссылки ведут на УТКО-копию `{i}_utko.jpg` (≤500 КБ) — см. докстринг модуля.
    """
    urls = [_abs_url(u, base_url) for u in (inc.photo_urls or [])]
    urls = [_utko_variant_url(u) for u in urls if u][:_PHOTO_SLOTS]
    return urls + [""] * (_PHOTO_SLOTS - len(urls))


def _subject(
    inc: Incident,
    region_by_mno: dict | None,
    region_index: dict | None = None,
) -> str:
    """Субъект РФ КАК В УТКО — из НАШЕГО справочника. Приоритет источников:

    1. МНО инцидента (region_by_mno: mno_id → Mno.region_code → Region.name) — самый
       авторитетный источник, МНО пришло из ФГИС вместе с регионом;
    2. текст inc.region, сопоставленный со справочником (region_index) → каноническое
       Region.name. Гасит косметику DaData/AI: «Санкт-Петербург» → «г. Санкт-Петербург»,
       «Респ Татарстан» → «Республика Татарстан», длинное тире → дефис;
    3. inc.region как есть — последний фолбэк, если субъекта нет в справочнике. Терять
       данные хуже: УТКО такую строку отвергнет, но оператор увидит, что чинить.
    """
    if region_by_mno and getattr(inc, "mno_id", None):
        name = region_by_mno.get(inc.mno_id)
        if name:
            return name
    if region_index:
        canonical = region_index.get(region_match_key(normalize_region(inc.region)))
        if canonical:
            return canonical
    return inc.region


def _region_code(
    inc: Incident,
    region_code_by_mno: dict | None,
    region_code_index: dict | None = None,
) -> str | None:
    """Код субъекта РФ инцидента (Region.code) — ЗЕРКАЛО _subject, но для пояса.

    Тем же приоритетом источников, что _subject резолвит ИМЯ субъекта, здесь резолвим
    КОД (нужен services/region_tz.region_utc_offset для пересчёта времени в UTC):
    1. МНО инцидента (region_code_by_mno: mno_id → Mno.region_code);
    2. текст inc.region, сопоставленный со справочником (region_code_index) → Region.code;
    3. None — субъект не распознан (region_utc_offset вернёт дефолт МСК +3).
    """
    if region_code_by_mno and getattr(inc, "mno_id", None):
        code = region_code_by_mno.get(inc.mno_id)
        if code:
            return code
    if region_code_index:
        code = region_code_index.get(region_match_key(normalize_region(inc.region)))
        if code:
            return code
    return None


def _row(
    inc: Incident,
    base_url: str,
    type_labels: dict | None,
    region_by_mno: dict | None,
    region_index: dict | None = None,
    region_code_by_mno: dict | None = None,
    region_code_index: dict | None = None,
) -> list:
    offset = region_utc_offset(_region_code(inc, region_code_by_mno, region_code_index))
    return [
        _subject(inc, region_by_mno, region_index),
        inc.mno_reg or "",
        _photo_datetime(inc, offset),
        _address(inc),
        _type_label(inc, type_labels),
        _subtype_label(inc.incident_subtype),  # Подтип (только у типа no_access; иначе "")
        inc.comment or "",
        *_photo_urls(inc, base_url),
    ]


def build_utko_xlsx(
    rows: Iterable[Incident],
    base_url: str = "",
    type_labels: dict | None = None,
    region_by_mno: dict | None = None,
    region_index: dict | None = None,
    region_code_by_mno: dict | None = None,
    region_code_index: dict | None = None,
) -> bytes:
    """Заполняет клиентский шаблон УТКО и возвращает .xlsx (bytes).

    Данные пишутся СТРОГО с 5-й строки листа «Реестр для инцидентов» — 4-строчная шапка
    шаблона не трогается. Ссылки на фото — просто текстовые URL.

    Колонка «Субъект РФ» (см. _subject) собирается из двух карт НАШЕГО справочника:
    region_by_mno — {mno_id: Region.name} для инцидентов с МНО;
    region_index  — {region_match_key(Region.name): Region.name} (region.canonical_index)
    для инцидентов БЕЗ МНО: приводит текст inc.region к канону справочника.
    Обе None → в файл уйдёт сырой inc.region (DaData-текст, УТКО его может не принять).

    Колонка «Дата и время фотофиксации» отдаётся в UTC (пересчёт настенного времени по
    поясу региона инцидента, см. _photo_datetime + services/region_tz). Пояс резолвится по
    КОДУ субъекта инцидента (ЗЕРКАЛО «Субъекта РФ», но код вместо имени, см. _region_code):
    region_code_by_mno — {mno_id: Mno.region_code} (mno.region_codes_by_mno);
    region_code_index  — {region_match_key(Region.name): Region.code} (region.code_index).
    Обе None → неизвестный пояс → дефолт МСК +3.
    """
    wb = load_workbook(_TEMPLATE_PATH)
    ws = wb[_DATA_SHEET]
    wb.active = wb.sheetnames.index(_DATA_SHEET)

    row_idx = _DATA_START_ROW
    for inc in rows:
        for col_idx, value in enumerate(
            _row(
                inc, base_url, type_labels, region_by_mno, region_index,
                region_code_by_mno, region_code_index,
            ),
            start=1,
        ):
            ws.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
